#!/usr/bin/env python3
"""Retro CLI - Batch trade submission and verification tool for CIT mocknet."""

import argparse
import csv
import os
import sys
import time
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import requests

# H2 database access via subprocess (no JVM dependency)
import subprocess


TERMINAL_STATUSES = {"NETTED", "SETTLED", "REJECTED"}
PASS_STATUSES = {"NETTED", "SETTLED"}
DEFAULT_URL = "http://localhost:8080"
DEFAULT_DB = "mocknet/data/coredb"
HEADERS = ["participant", "payload", "response", "result"]


def parse_args():
    parser = argparse.ArgumentParser(
        prog="retro",
        description="Batch trade submission and verification tool for CIT mocknet.",
    )
    parser.add_argument("input_file", help="Path to .xlsx or .csv file")
    parser.add_argument(
        "--url",
        default=os.environ.get("RETRO_MOCKNET_URL", DEFAULT_URL),
        help=f"Mocknet base URL (default: {DEFAULT_URL})",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB,
        help=f"H2 database file path without extension (default: {DEFAULT_DB})",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite input file instead of creating _results file",
    )
    parser.add_argument(
        "--output",
        help="Write results to a specific .csv or .xlsx file",
    )
    parser.add_argument(
        "--no-header",
        action="store_true",
        help="Treat columns as positional: A=participant, B=payload, C=response, D=result",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="Per-trade DB verification timeout in seconds (default: 30)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="Delay between submissions in seconds (default: 0.5)",
    )
    parser.add_argument(
        "--skip-db",
        action="store_true",
        help="Skip DB verification, only check HTTP response",
    )
    return parser.parse_args()


def extract_trade_id(xml_payload):
    """Extract tradeId from XML payload."""
    try:
        root = ET.fromstring(xml_payload)
        trade_id_elem = root.find(".//tradeId")
        if trade_id_elem is not None and trade_id_elem.text:
            return trade_id_elem.text.strip()
    except ET.ParseError:
        pass
    return None


def submit_trade(url, xml_payload):
    """POST XML payload to mocknet. Returns (status_code, response_body)."""
    try:
        resp = requests.post(
            f"{url}/api/trades",
            data=xml_payload.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=10,
        )
        return resp.status_code, resp.text
    except requests.ConnectionError:
        return None, "Connection refused - is mocknet running?"
    except requests.Timeout:
        return None, "Request timed out"
    except requests.RequestException as e:
        return None, str(e)


def query_trade_status_via_api(url, trade_id):
    """Query trade status via REST API as fallback."""
    try:
        resp = requests.get(f"{url}/api/trades", timeout=5)
        if resp.status_code == 200:
            trades = resp.json()
            for trade in trades:
                if trade.get("tradeId") == trade_id:
                    return trade.get("status")
    except Exception:
        pass
    return None


def find_h2_jar():
    """Find H2 jar in mocknet's Maven dependencies."""
    m2_h2 = Path.home() / ".m2" / "repository" / "com" / "h2database" / "h2"
    if m2_h2.exists():
        for version_dir in sorted(m2_h2.iterdir(), reverse=True):
            jar = version_dir / f"h2-{version_dir.name}.jar"
            if jar.exists():
                return str(jar)
    return None


def query_trade_status_h2(db_path, trade_id, h2_jar):
    """Query trade status directly from H2 database via subprocess."""
    if not h2_jar:
        return None

    db_file = Path(db_path)
    if not db_file.with_suffix(".mv.db").exists():
        return None

    jdbc_url = f"jdbc:h2:file:{db_path};DB_CLOSE_ON_EXIT=FALSE;AUTO_RECONNECT=TRUE;AUTO_SERVER=TRUE;IFEXISTS=TRUE"
    sql = f"SELECT STATUS FROM TRADES WHERE TRADE_ID = '{trade_id}';"

    try:
        result = subprocess.run(
            [
                "java",
                "-cp",
                h2_jar,
                "org.h2.tools.Shell",
                "-url",
                jdbc_url,
                "-user",
                "sa",
                "-password",
                "",
                "-sql",
                sql,
            ],
            capture_output=True,
            text=True,
            timeout=5,
        )
        output = result.stdout.strip()
        for line in output.splitlines():
            line = line.strip()
            if line in (
                "RECEIVED",
                "VALIDATED",
                "MATCHED",
                "NETTED",
                "SETTLED",
                "REJECTED",
            ):
                return line
    except (subprocess.TimeoutExpired, FileNotFoundError):
        pass
    return None


def poll_trade_status(url, db_path, trade_id, timeout, h2_jar, skip_db):
    """Poll for trade terminal status. Tries H2 DB first, falls back to REST API."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        status = None
        if not skip_db and h2_jar:
            status = query_trade_status_h2(db_path, trade_id, h2_jar)
        if status is None:
            status = query_trade_status_via_api(url, trade_id)

        if status in TERMINAL_STATUSES:
            return status
        time.sleep(1)
    return "TIMEOUT"


def process_trade(row, url, db_path, timeout, h2_jar, skip_db):
    """Process a single trade row. Returns (response_text, result)."""
    participant = row.get("participant", "").strip()
    payload = row.get("payload", "").strip()

    if not payload:
        return "No payload", "SKIP"

    trade_id = extract_trade_id(payload)
    status_code, response_body = submit_trade(url, payload)

    if status_code is None:
        return response_body, "FAIL"

    if status_code != 202:
        return f"HTTP {status_code}: {response_body}", "FAIL"

    # HTTP was accepted - now verify DB status
    if trade_id:
        db_status = poll_trade_status(url, db_path, trade_id, timeout, h2_jar, skip_db)
        if db_status == "TIMEOUT":
            return f"HTTP 202 Accepted | DB: TIMEOUT", "FAIL"
        elif db_status in PASS_STATUSES:
            return f"HTTP 202 Accepted | DB: {db_status}", "PASS"
        else:
            return f"HTTP 202 Accepted | DB: {db_status}", "FAIL"
    else:
        # Can't extract trade ID - just report HTTP success
        return f"HTTP 202 Accepted | DB: N/A (no tradeId)", "PASS"


# --- File I/O ---


def read_csv(filepath, no_header):
    """Read CSV file, return list of row dicts."""
    rows = []
    with open(filepath, newline="", encoding="utf-8") as f:
        if no_header:
            reader = csv.reader(f)
            for r in reader:
                if len(r) >= 2:
                    rows.append(
                        {
                            "participant": r[0],
                            "payload": r[1],
                            "response": r[2] if len(r) > 2 else "",
                            "result": r[3] if len(r) > 3 else "",
                        }
                    )
        else:
            reader = csv.DictReader(f)
            found = {h.lower().strip() for h in (reader.fieldnames or [])}
            if "participant" not in found or "payload" not in found:
                print(
                    f"Error: CSV must have 'participant' and 'payload' headers. Found: {reader.fieldnames}",
                    file=sys.stderr,
                )
                sys.exit(1)
            for r in reader:
                normalized = {k.lower().strip(): v for k, v in r.items()}
                rows.append(
                    {
                        "participant": normalized.get("participant", ""),
                        "payload": normalized.get("payload", ""),
                        "response": normalized.get("response", ""),
                        "result": normalized.get("result", ""),
                    }
                )
    return rows


def write_csv(filepath, rows, no_header):
    """Write rows back to CSV."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        if no_header:
            writer = csv.writer(f)
            for r in rows:
                writer.writerow(
                    [r["participant"], r["payload"], r["response"], r["result"]]
                )
        else:
            writer = csv.DictWriter(f, fieldnames=HEADERS)
            writer.writeheader()
            for r in rows:
                writer.writerow({h: r[h] for h in HEADERS})


def read_xlsx(filepath, no_header):
    """Read Excel file, return (workbook, rows list)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = []

    if no_header:
        for row in ws.iter_rows(min_row=1, values_only=False):
            cells = [c.value or "" for c in row]
            participant = str(cells[0]) if len(cells) > 0 else ""
            payload = str(cells[1]) if len(cells) > 1 else ""
            if participant.strip() or payload.strip():
                rows.append(
                    {
                        "participant": participant,
                        "payload": payload,
                        "response": str(cells[2]) if len(cells) > 2 else "",
                        "result": str(cells[3]) if len(cells) > 3 else "",
                        "_row_num": row[0].row,
                    }
                )
    else:
        header_row = [str(c.value or "").lower().strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "participant" not in header_row or "payload" not in header_row:
            print(
                f"Error: Excel must have 'participant' and 'payload' headers. Found: {header_row}",
                file=sys.stderr,
            )
            sys.exit(1)

        col_map = {name: idx for idx, name in enumerate(header_row)}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = [c.value or "" for c in row]
            participant = str(cells[col_map["participant"]]) if "participant" in col_map else ""
            payload = str(cells[col_map["payload"]]) if "payload" in col_map else ""
            if participant.strip() or payload.strip():
                rows.append(
                    {
                        "participant": participant,
                        "payload": payload,
                        "response": "",
                        "result": "",
                        "_row_num": row[0].row,
                        "_col_response": col_map.get("response"),
                        "_col_result": col_map.get("result"),
                    }
                )

    return wb, rows


def write_xlsx(wb, rows, filepath, no_header):
    """Write results back to Excel workbook."""
    ws = wb.active

    if no_header:
        for r in rows:
            row_num = r["_row_num"]
            ws.cell(row=row_num, column=3, value=r["response"])
            ws.cell(row=row_num, column=4, value=r["result"])
    else:
        # Ensure response and result columns exist in header
        header_row = [str(c.value or "").lower().strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {name: idx + 1 for idx, name in enumerate(header_row)}  # 1-based

        if "response" not in col_map:
            col_map["response"] = ws.max_column + 1
            ws.cell(row=1, column=col_map["response"], value="response")
        if "result" not in col_map:
            col_map["result"] = max(ws.max_column, col_map["response"]) + 1
            ws.cell(row=1, column=col_map["result"], value="result")

        for r in rows:
            row_num = r["_row_num"]
            ws.cell(row=row_num, column=col_map["response"], value=r["response"])
            ws.cell(row=row_num, column=col_map["result"], value=r["result"])

    wb.save(filepath)


def output_path(input_file, in_place, output_file=None):
    """Determine output file path."""
    if output_file:
        return str(Path(output_file))
    if in_place:
        return input_file
    p = Path(input_file)
    return str(p.with_stem(p.stem + "_results"))


def main():
    args = parse_args()
    input_file = args.input_file
    ext = Path(input_file).suffix.lower()

    if ext not in (".csv", ".xlsx"):
        print(f"Error: Expected .xlsx or .csv file, got '{ext}'", file=sys.stderr)
        sys.exit(1)

    if not Path(input_file).exists():
        print(f"Error: File not found: {input_file}", file=sys.stderr)
        sys.exit(1)

    # Resolve DB path
    db_path = str(Path(args.db).resolve()) if args.db else None
    # Strip .mv.db extension if provided
    if db_path and db_path.endswith(".mv.db"):
        db_path = db_path[:-6]

    # Find H2 jar for direct DB access
    h2_jar = None
    if not args.skip_db:
        h2_jar = find_h2_jar()
        if h2_jar:
            print(f"H2 jar found: {h2_jar}")
        else:
            print("Warning: H2 jar not found in ~/.m2. Using REST API for status verification.")
            print("  (Run 'cd mocknet && mvn dependency:resolve' to download H2 jar)")

    # Check DB file exists
    if not args.skip_db and db_path:
        db_file = Path(db_path).with_suffix(".mv.db")
        if not db_file.exists():
            print(f"Warning: Database not found at {db_file}. DB verification will use REST API.")
            h2_jar = None

    # Read input
    if ext == ".csv":
        rows = read_csv(input_file, args.no_header)
        wb = None
    else:
        wb, rows = read_xlsx(input_file, args.no_header)

    if not rows:
        print("No data rows found in input file.", file=sys.stderr)
        sys.exit(1)

    total = len(rows)
    pass_count = 0
    fail_count = 0
    skip_count = 0

    print(f"\nProcessing {total} trades -> {args.url}")
    print("-" * 60)

    pending_verification = []

    for i, row in enumerate(rows):
        participant = row.get("participant", "unknown").strip()
        payload = row.get("payload", "").strip()

        if not payload:
            row["response"] = "No payload"
            row["result"] = "SKIP"
            skip_count += 1
            print(f"  [{i+1}/{total}] {participant} -> \033[33mSKIP\033[0m ({row['response']})")
            continue

        trade_id = extract_trade_id(payload)
        status_code, response_body = submit_trade(args.url, payload)

        if status_code is None:
            row["response"] = response_body
            row["result"] = "FAIL"
            fail_count += 1
            print(f"  [{i+1}/{total}] {participant} -> \033[31mFAIL\033[0m ({row['response']})")
            continue

        if status_code != 202:
            row["response"] = f"HTTP {status_code}: {response_body}"
            row["result"] = "FAIL"
            fail_count += 1
            print(f"  [{i+1}/{total}] {participant} -> \033[31mFAIL\033[0m ({row['response']})")
            continue

        if trade_id and not args.skip_db and h2_jar:
            pending_verification.append((i, trade_id))
            print(f"  [{i+1}/{total}] {participant} -> \033[34mQUEUED\033[0m (HTTP 202 Accepted, awaiting DB)")
        else:
            row["response"] = (
                "HTTP 202 Accepted | DB: N/A"
                + (" (no tradeId)" if not trade_id else " (no DB)")
            )
            row["result"] = "PASS"
            pass_count += 1
            print(f"  [{i+1}/{total}] {participant} -> \033[32mPASS\033[0m ({row['response']})")

        if i < total - 1:
            time.sleep(args.delay)

    for row_idx, trade_id in pending_verification:
        row = rows[row_idx]
        participant = row.get("participant", "unknown").strip()
        db_status = poll_trade_status(
            args.url, db_path, trade_id, args.timeout, h2_jar, args.skip_db
        )
        if db_status == "TIMEOUT":
            row["response"] = "HTTP 202 Accepted | DB: TIMEOUT"
            row["result"] = "FAIL"
            fail_count += 1
            symbol = "\033[31mFAIL\033[0m"
        elif db_status in PASS_STATUSES:
            row["response"] = f"HTTP 202 Accepted | DB: {db_status}"
            row["result"] = "PASS"
            pass_count += 1
            symbol = "\033[32mPASS\033[0m"
        else:
            row["response"] = f"HTTP 202 Accepted | DB: {db_status}"
            row["result"] = "FAIL"
            fail_count += 1
            symbol = "\033[31mFAIL\033[0m"

        print(f"  [verify] {participant} -> {symbol} ({row['response']})")

    # Write output
    out_file = output_path(input_file, args.in_place, args.output)
    out_ext = Path(out_file).suffix.lower()
    if out_ext not in (".csv", ".xlsx"):
        print(
            f"Error: Output file must end in .csv or .xlsx, got '{out_ext}'",
            file=sys.stderr,
        )
        sys.exit(1)

    if out_ext == ".csv":
        write_csv(out_file, rows, args.no_header)
    else:
        if wb is None:
            fresh_wb = openpyxl.Workbook()
            fresh_ws = fresh_wb.active
            fresh_ws.title = "trades"
            if args.no_header:
                for row_idx, row in enumerate(rows, start=1):
                    fresh_ws.cell(row=row_idx, column=1, value=row["participant"])
                    fresh_ws.cell(row=row_idx, column=2, value=row["payload"])
                    fresh_ws.cell(row=row_idx, column=3, value=row["response"])
                    fresh_ws.cell(row=row_idx, column=4, value=row["result"])
            else:
                fresh_ws.append(HEADERS)
                for row in rows:
                    fresh_ws.append([row[h] for h in HEADERS])
            fresh_wb.save(out_file)
        else:
            write_xlsx(wb, rows, out_file, args.no_header)

    print("-" * 60)
    print(f"Results: {pass_count} PASS, {fail_count} FAIL, {skip_count} SKIP out of {total} trades")
    print(f"Written to: {out_file}")


if __name__ == "__main__":
    main()
